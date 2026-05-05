import json
import shutil
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE         = Path(__file__).parent
INBODY_DIR   = BASE.parent / "Inbody"
DATA_FILE    = INBODY_DIR / "inbody_data.json"
OUTPUT_FILE  = INBODY_DIR / "InBody_數據報表.xlsx"
INPUT_DIR    = INBODY_DIR / "要新增的數據"
PROCESSED_DIR = INBODY_DIR / "已處理"

INPUT_DIR.mkdir(exist_ok=True)
PROCESSED_DIR.mkdir(exist_ok=True)

with open(DATA_FILE, encoding="utf-8") as f:
    data = json.load(f)

records = sorted(data["records"], key=lambda r: r["date"])
history = sorted(data["history"], key=lambda r: r["date"])

wb = Workbook()

border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
purple_fill = PatternFill("solid", fgColor="6B3FA0")
light_purple = PatternFill("solid", fgColor="F3ECFC")
metric_fill = PatternFill("solid", fgColor="EDE7F6")
white_font = Font(bold=True, color="FFFFFF", size=11)


def hcell(ws, row, col, val, fill, font=None):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = fill
    c.font = font or white_font
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border
    return c


def dcell(ws, row, col, val, alt=False):
    c = ws.cell(row=row, column=col, value=val)
    c.alignment = Alignment(horizontal="center")
    c.border = border
    if alt:
        c.fill = light_purple
    return c


# ── 量測摘要（轉置：指標固定在第一欄，日期橫向展開）─────────────
ws1 = wb.active
ws1.title = "量測摘要"

METRICS = [
    "體重(kg)", "骨骼肌重SMM(kg)", "體脂肪重(kg)", "BMI(kg/m²)", "體脂肪率(%)",
    "除脂體重(kg)", "身體總水量(L)", "蛋白質重量(kg)", "礦物質重量(kg)", "InBody評分",
    "目標體重(kg)", "體重控制(kg)", "脂肪控制(kg)", "肌肉控制(kg)",
    "內臟脂肪級別", "腰臀圍比", "SMI肌肉量指數(kg/m²)", "骨礦含量(kg)",
    "基礎代謝率(kcal)", "建議熱量攝取(kcal)",
]

dates = [r["date"] for r in records]

hcell(ws1, 1, 1, "指標", purple_fill)
for col_idx, d in enumerate(dates, 2):
    hcell(ws1, 1, col_idx, d, purple_fill)

for row_idx, metric in enumerate(METRICS, 2):
    c = ws1.cell(row=row_idx, column=1, value=metric)
    c.font = Font(bold=True)
    c.fill = metric_fill
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = border
    alt = row_idx % 2 == 0
    for col_idx, rec in enumerate(records, 2):
        dcell(ws1, row_idx, col_idx, rec.get(metric), alt)

ws1.column_dimensions["A"].width = 24
for col_idx in range(2, len(dates) + 2):
    ws1.column_dimensions[get_column_letter(col_idx)].width = 16
ws1.row_dimensions[1].height = 22


# ── 部位別分析 ────────────────────────────────────────────────
ws2 = wb.create_sheet("部位別分析")
PARTS = ["右上肢", "左上肢", "軀幹", "右下肢", "左下肢"]
blue_fill = PatternFill("solid", fgColor="4472C4")
red_fill = PatternFill("solid", fgColor="C0504D")
blue_alt = PatternFill("solid", fgColor="DCE6F1")

row = 1
ws2.cell(row=row, column=1, value="部位別肌肉分析").font = Font(bold=True, size=12)
row += 1

muscle_headers = ["部位"] + [f"{d} 肌肉(kg)" for d in dates] + [f"{d} %" for d in dates]
for ci, h in enumerate(muscle_headers, 1):
    hcell(ws2, row, ci, h, blue_fill)
row += 1

for pi, part in enumerate(PARTS):
    alt = pi % 2 == 0
    ws2.cell(row=row, column=1, value=part).border = border
    if alt:
        ws2.cell(row=row, column=1).fill = blue_alt
    col = 2
    for rec in records:
        c = ws2.cell(row=row, column=col, value=rec["部位肌肉"].get(f"{part}_kg"))
        c.border = border
        c.alignment = Alignment(horizontal="center")
        if alt:
            c.fill = blue_alt
        col += 1
    for rec in records:
        c = ws2.cell(row=row, column=col, value=rec["部位肌肉"].get(f"{part}_%"))
        c.border = border
        c.alignment = Alignment(horizontal="center")
        if alt:
            c.fill = blue_alt
        col += 1
    row += 1

row += 1
ws2.cell(row=row, column=1, value="部位別脂肪分析").font = Font(bold=True, size=12)
row += 1

fat_headers = ["部位"] + [f"{d} 脂肪(kg)" for d in dates] + [f"{d} %" for d in dates]
for ci, h in enumerate(fat_headers, 1):
    hcell(ws2, row, ci, h, red_fill)
row += 1

for pi, part in enumerate(PARTS):
    alt = pi % 2 == 0
    ws2.cell(row=row, column=1, value=part).border = border
    if alt:
        ws2.cell(row=row, column=1).fill = PatternFill("solid", fgColor="F2DCDB")
    col = 2
    for rec in records:
        c = ws2.cell(row=row, column=col, value=rec["部位脂肪"].get(f"{part}_kg"))
        c.border = border
        c.alignment = Alignment(horizontal="center")
        if alt:
            c.fill = PatternFill("solid", fgColor="F2DCDB")
        col += 1
    for rec in records:
        c = ws2.cell(row=row, column=col, value=rec["部位脂肪"].get(f"{part}_%"))
        c.border = border
        c.alignment = Alignment(horizontal="center")
        if alt:
            c.fill = PatternFill("solid", fgColor="F2DCDB")
        col += 1
    row += 1

for col in ws2.columns:
    max_len = max(len(str(c.value)) if c.value else 0 for c in col)
    ws2.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 4, 14)


# ── 部位別圍度 ────────────────────────────────────────────────
ws3 = wb.create_sheet("部位別圍度")
green_fill = PatternFill("solid", fgColor="375623")
green_alt = PatternFill("solid", fgColor="EBF1DE")
CIRC_PARTS = ["頸部", "胸部", "腰部", "臀部", "右臂", "左臂", "右大腿", "左大腿"]

circ_headers = ["部位"] + [f"{d} (cm)" for d in dates]
if len(dates) >= 2:
    circ_headers.append("最新變化 (cm)")
for ci, h in enumerate(circ_headers, 1):
    hcell(ws3, 1, ci, h, green_fill)

for ri, part in enumerate(CIRC_PARTS, 2):
    alt = ri % 2 == 0
    c = ws3.cell(row=ri, column=1, value=part)
    c.border = border
    c.alignment = Alignment(horizontal="center")
    if alt:
        c.fill = green_alt
    for ci, rec in enumerate(records, 2):
        val = rec["圍度(cm)"].get(part)
        c = ws3.cell(row=ri, column=ci, value=val)
        c.border = border
        c.alignment = Alignment(horizontal="center")
        if alt:
            c.fill = green_alt
    if len(dates) >= 2:
        v1 = records[-2]["圍度(cm)"].get(part) or 0
        v2 = records[-1]["圍度(cm)"].get(part) or 0
        diff = round(v2 - v1, 1)
        c = ws3.cell(row=ri, column=len(dates) + 2, value=diff)
        c.border = border
        c.alignment = Alignment(horizontal="center")
        c.font = Font(color="C0504D" if diff > 0 else ("375623" if diff < 0 else "000000"), bold=True)
        if alt:
            c.fill = green_alt

for col in ws3.columns:
    max_len = max(len(str(c.value)) if c.value else 0 for c in col)
    ws3.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 4, 16)


# ── 歷程紀錄 ──────────────────────────────────────────────────
ws4 = wb.create_sheet("歷程紀錄")
hist_headers = ["日期", "體重(kg)", "骨骼肌重(kg)", "體脂肪率(%)"]
for ci, h in enumerate(hist_headers, 1):
    hcell(ws4, 1, ci, h, purple_fill)

for ri, rec in enumerate(history, 2):
    alt = ri % 2 == 0
    for ci, key in enumerate(["date", "體重(kg)", "骨骼肌重(kg)", "體脂肪率(%)"], 1):
        c = ws4.cell(row=ri, column=ci, value=rec.get(key))
        c.alignment = Alignment(horizontal="center")
        c.border = border
        if alt:
            c.fill = light_purple

for col in ws4.columns:
    max_len = max(len(str(c.value)) if c.value else 0 for c in col)
    ws4.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 4, 16)


wb.save(OUTPUT_FILE)
print(f"Excel 已儲存：{OUTPUT_FILE}")

if "--move" in sys.argv:
    moved = []
    for f in INPUT_DIR.iterdir():
        if f.suffix.lower() in {".jpg", ".jpeg", ".png"}:
            shutil.move(str(f), PROCESSED_DIR / f.name)
            moved.append(f.name)
    if moved:
        print(f"已移動 {len(moved)} 張圖片至：{PROCESSED_DIR}")
